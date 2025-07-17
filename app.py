#!/usr/bin/env python3

import os
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# --- Config handling ----------------------------------------------------------------

def load_config(config_path='config.json'):
    """
    Load the list of rows to remove from a JSON config file. If the file does not exist,
    create it with default values and prompt the user to review.
    """
    default_rows = [
        "Timestamp",
        "Time Since Page Load",
        "Initiator",
        "frame",
        "hitId",
        "isMultiSuiteTagging",
        "isTruncated",
        "reportSuiteIds",
        "returnType",
        "trackingServer",
        "version",
        ".a",
        ".activitymap",
        ".c",
        "a.",
        "Activity Map Link",
        "Activity Map Page",
        "Activity Map Page Type",
        "Activity Map Region",
        "activitymap.",
        "Audience Manager Blob",
        "Audience Manager Location Hint",
        "Browser Window Height",
        "Browser Window Width",
        "c.getPreviousValue",
        "c.getQueryParam",
        "c.pt",
        "Character Set",
        "ClickMap Object ID",
        "ClickMap Object Tag Name",
        "ClickMap Page ID",
        "ClickMap Page ID Type",
        "Color quality",
        "Context Data",
        "Cookie Domain",
        "Cookies Enabled",
        "Currency Code"
    ]
    if not os.path.exists(config_path):
        with open(config_path, 'w') as f:
            json.dump(default_rows, f, indent=4)
        raise FileNotFoundError(f"No config found. A default {config_path} has been created. Please review and run again.")
    with open(config_path) as f:
        return json.load(f)


# --- Core comparison logic ----------------------------------------------------------

def load_and_clean(path, rows_to_remove):
    """
    Load an XLSX file, set the first column as the index of attributes,
    drop the header row and any rows listed in rows_to_remove.
    """
    df = pd.read_excel(path, header=None)
    df = df.set_index(0)
    df.index.name = None
    df = df.drop('Solution', errors='ignore')
    df = df[~df.index.isin(rows_to_remove)]
    return df


def compare_and_write(prod_path, dev_path, rows_to_remove, output_path):
    """
    Compare cleaned production and development DataFrames, write two sheets,
    and create a Differences sheet with highlights for changed cells.
    """
    prod_df = load_and_clean(prod_path, rows_to_remove)
    dev_df = load_and_clean(dev_path, rows_to_remove)
    # If number of columns differ, take only first and last analytics columns
    if prod_df.shape[1] != dev_df.shape[1]:
        prod_df = prod_df[[prod_df.columns[0], prod_df.columns[-1]]]
        dev_df = dev_df[[dev_df.columns[0], dev_df.columns[-1]]]

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Write production and development to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        prod_df.to_excel(writer, sheet_name='Production')
        dev_df.to_excel(writer, sheet_name='Development')

    # Load workbook to create Differences sheet
    wb = load_workbook(output_path)
    ws_prod = wb['Production']
    ws_dev = wb['Development']
    ws_diff = wb.create_sheet('Differences')

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Copy production values and highlight differences
    for row in ws_prod.iter_rows(min_row=1, max_row=ws_prod.max_row, min_col=1, max_col=ws_prod.max_column):
        for cell in row:
            diff_cell = ws_diff.cell(row=cell.row, column=cell.column, value=cell.value)
            dev_value = ws_dev.cell(row=cell.row, column=cell.column).value
            if cell.value != dev_value:
                diff_cell.fill = yellow_fill

    wb.save(output_path)


# --- GUI ------------------------------------------------------------------------------
class App:
    """
    Tkinter GUI for selecting two XLSX files and running the comparison.
    """
    def __init__(self, root):
        self.root = root
        self.root.title("AEP XLSX Comparator")
        self.create_widgets()

    def create_widgets(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.grid(row=0, column=0)

        # Platform input
        ttk.Label(frame, text="Platform:").grid(row=0, column=0, sticky='w')
        self.platform_entry = ttk.Entry(frame, width=50)
        self.platform_entry.grid(row=0, column=1, padx=5, columnspan=1)

        # Production file input
        ttk.Label(frame, text="Production XLSX:").grid(row=1, column=0, sticky='w')
        self.prod_entry = ttk.Entry(frame, width=50)
        self.prod_entry.grid(row=1, column=1, padx=5)
        ttk.Button(frame, text="Browse", command=self.browse_prod).grid(row=1, column=2)

        # Development file input
        ttk.Label(frame, text="Development XLSX:").grid(row=2, column=0, sticky='w')
        self.dev_entry = ttk.Entry(frame, width=50)
        self.dev_entry.grid(row=2, column=1, padx=5)
        ttk.Button(frame, text="Browse", command=self.browse_dev).grid(row=2, column=2)

        # Progress bar
        self.progress = ttk.Progressbar(frame, orient='horizontal', mode='determinate', length=400)
        self.progress.grid(row=3, column=0, columnspan=3, pady=10)

        # Execute button
        self.go_button = ttk.Button(frame, text="Go", command=self.run)
        self.go_button.grid(row=4, column=1)

    def browse_prod(self):
        path = filedialog.askopenfilename(filetypes=[('Excel files','*.xlsx')])
        if path:
            self.prod_entry.delete(0, tk.END)
            self.prod_entry.insert(0, path)

    def browse_dev(self):
        path = filedialog.askopenfilename(filetypes=[('Excel files','*.xlsx')])
        if path:
            self.dev_entry.delete(0, tk.END)
            self.dev_entry.insert(0, path)

    def run(self):
        platform = self.platform_entry.get().strip()
        prod = self.prod_entry.get()
        dev = self.dev_entry.get()
        if not prod or not dev:
            messagebox.showerror("Error", "Please select both production and development files.")
            return

        self.go_button.config(state='disabled')
        self.progress['value'] = 10
        self.root.update_idletasks()

        try:
            rows_to_remove = load_config()
        except FileNotFoundError as e:
            messagebox.showinfo("Config", str(e))
            self.go_button.config(state='normal')
            self.progress['value'] = 0
            return

        self.progress['value'] = 30
        self.root.update_idletasks()

        # Construct output file path
        today = datetime.datetime.today().strftime('%d_%m_%y')
        safe_platform = platform if platform else 'platform'
        filename = f"{safe_platform}_comparison_output_{today}.xlsx"
        output_dir = 'output'
        output_path = os.path.join(output_dir, filename)

        # Execute comparison
        compare_and_write(prod, dev, rows_to_remove, output_path)

        self.progress['value'] = 100
        self.root.update_idletasks()
        messagebox.showinfo("Complete", f"Comparison complete. Output saved as {filename}")

        self.go_button.config(state='normal')
        self.progress['value'] = 0


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()